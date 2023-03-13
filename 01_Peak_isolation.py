import openpyxl
import datetime

starttime = datetime.datetime.now()

importfile_name = ('long-term enrichment_1M_01.xlsx','long-term enrichment_1M_02.xlsx')
nozero_name = ('nozero_long-term enrichment_1M_01.xlsx', 'nozero_long-term enrichment_1M_02.xlsx')
outputfile_name = ('peak_long-term enrichment_1M_01.xlsx','peak_long-term enrichment_1M_02.xlsx')

for i in range(len(importfile_name)):
    ip_name = importfile_name[i]
    nz_name = nozero_name[i]
    # op_name = outputfile_name[i]

    ipwb = openpyxl.load_workbook(ip_name)
    ipws = ipwb.active
    nrows = ipws.max_row
    ncols = ipws.max_column

    nzwb = openpyxl.Workbook()
    nzws = nzwb.active

    print('for wb %s, the row number is %d, and the column number is %d' % (ip_name, nrows, ncols))

    group_num = int(ncols / 4)
    for col_num in range(group_num):
        col_ms = col_num * 4 + 1
        col_in = col_num * 4 + 2

        # copy the name and title for each raw data (row 1-9)
        for row_num in range(1, 10):
            nzws.cell(row_num, col_ms).value = ipws.cell(row_num, col_ms).value
            nzws.cell(row_num, col_in).value = ipws.cell(row_num, col_in).value

        #  remove all the O point from the raw data
        nz_row = 10
        for row_num in range(10, nrows + 1):
            ms = ipws.cell(row_num, col_ms).value
            intensity = ipws.cell(row_num, col_in).value
            if isinstance(ms, float) or isinstance(ms, int):
                if intensity > 0:
                    nzws.cell(nz_row, col_ms).value = ms
                    nzws.cell(nz_row, col_in).value = intensity
                    nz_row += 1
                else:
                    continue
            else:
                break
    nzwb.save(nz_name)
    print('the zero removal for %s is finished' % nz_name)
print('the zero removal for all workbooks are finished')

# add a column = (A2-A1)*(A3-A2) and select the peak point
for i in range(len(importfile_name)):
    nz_name = nozero_name[i]
    op_name = outputfile_name[i]

    ipwb = openpyxl.load_workbook(nz_name)
    ipws = ipwb.active
    nrows = ipws.max_row
    ncols = ipws.max_column

    opwb = openpyxl.Workbook()
    opws = opwb.active

    print('for wb %s, the row number is %d, and the column number is %d' % (nz_name, nrows, ncols))

    group_num = int((ncols + 2) / 4)
    for col_num in range(group_num):
        col_ms = col_num * 4 + 1
        col_in = col_num * 4 + 2

        # copy the name and title for each raw data (row 1-9)
        for row_num in range(1, 10):
            opws.cell(row_num, col_ms).value = ipws.cell(row_num, col_ms).value
            opws.cell(row_num, col_in).value = ipws.cell(row_num, col_in).value
        opws.cell(9, col_in + 1).value = 'multifactor'

        #  add a column of the multiple factor
        op_row = 10
        for row_num in range(11, nrows):
            ms_2 = ipws.cell(row_num, col_ms).value
            ms_3 = ipws.cell(row_num + 1, col_ms).value
            intensity_1 = ipws.cell(row_num - 1, col_in).value
            intensity_2 = ipws.cell(row_num, col_in).value
            intensity_3 = ipws.cell(row_num + 1, col_in).value
            if isinstance(ms_3, float) or isinstance(ms_3, int):
                multifactor = (intensity_2 - intensity_1) * (intensity_3 - intensity_2)
                ipws.cell(row_num, col_in + 1).value = multifactor
                if multifactor < 0:
                    opws.cell(op_row, col_ms).value = ms_2
                    opws.cell(op_row, col_in).value = intensity_2
                    op_row += 1
                else:
                    continue
            else:
                break
    ipwb.save(nz_name)
    opwb.save(op_name)
    print('the peak isolation for workbook %s is done' % op_name)
print('the peak isolation for all workbooks are finished')

endtime = datetime.datetime.now()
runtime = float((endtime - starttime).seconds / 60)
print('the overall runing time is ', runtime)
