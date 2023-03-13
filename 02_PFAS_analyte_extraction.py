import openpyxl
import datetime

starttime = datetime.datetime.now()

importfile_name = ('PEAKS_enrichment_3M_data.xlsx', 'PEAKS_enrichment_5M_data.xlsx')
outpput1_name = ('PFAS_extract_enrichment_3M_data.xlsx', 'PFAS_extract_enrichment_5M_data.xlsx')
# output2_name = ('PFAS_quanti_L.xlsx','PFAS_quanti_R.xlsx','PFAS_quanti_P.xlsx','PFAS_quanti_control.xlsx','PFAS_quanti_AFFF.xlsx')
PFAS_analyate = 'PFAS_analyte searching list.xlsx'

# extract all PFAS analyate intensity
searchwb = openpyxl.load_workbook(PFAS_analyate)
searchws = searchwb.active
n_PFAS = searchws.max_row
masserror = 0.000005

for i in range (len(importfile_name)):
    ip_name = importfile_name[i]
    ipwb = openpyxl.load_workbook(ip_name)
    ipws = ipwb.active
    opwb = openpyxl.Workbook()
    opws = opwb.active

    ncols = ipws.max_column
    nrows = ipws.max_row
    # copy the analyte to the opws
    for m in range(1,4):
        for n in range (2, n_PFAS+1):
            opws.cell(n,m).value = searchws.cell(n,m).value
    print('the copy of PFAS analytes list is finished')

    # screening the target PFAS analytes from the searching list
    group_num = int ((ncols + 2)/4)
    # the column number from the import data
    for col_num in range (group_num):
        col_ms = col_num*4 + 1
        col_inten = col_num*4 +2
        op_inten = (col_num+1)*4 + 1
        op_RI = (col_num + 1)*4 +2
        op_conc = (col_num +1)*4 +3
        # name each column
        sample_id = ipws.cell(3,col_ms).value
        opws.cell(1,op_inten).value = sample_id
        opws.cell(2,op_inten).value = 'intensity'
        opws.cell(2, op_RI).value = 'RI'
        opws.cell(2, op_conc).value = 'concentration in ppb'

        #  get the m/z for target PFAS compounds
        for pfas_row in range(3,n_PFAS+1):
            PFAS_ms = searchws.cell(pfas_row,3).value
            # searching the target pfas m/z in each sample column
            checking = 0
            for row_num in range (10,nrows+1):
                sample_ms = ipws.cell(row_num, col_ms).value
                sample_in = ipws.cell(row_num, col_inten).value
                if isinstance(sample_ms, float) or isinstance(sample_ms, int):
                    if abs(1-sample_ms/PFAS_ms) < masserror:
                        opws.cell(pfas_row,op_inten).value = ipws.cell(row_num, col_inten).value
                        checking += 1
                        break
                    else:
                        continue
                else:
                    break
            if checking == 0:
                opws.cell(pfas_row,op_inten).value = 'not detected'
        print('the target PFAS screening is finished for sample', sample_id)
    opwb.save(outpput1_name[i])
    print('the target pfas screening is finished for workbook', outpput1_name[i])
endtime1 = datetime.datetime.now()
runtime1 = float ((endtime1-starttime).seconds/60)
print('the target pfas extraction is finished for all workbooks',' the overall running time is ', runtime1)


